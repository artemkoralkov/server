import re
from typing import Dict

from openpyxl.utils import get_column_letter


from openpyxl.reader.excel import load_workbook


def split(a, n):

    k, m = divmod(len(a), n)
    return (a[i * k + min(i, m):(i + 1) * k + min(i + 1, m)] for i in range(n))


def lesson_to_dict(lesson: str) -> Dict:
    """
    Keyword arguments:
    argument -- description
    Return: return_description
    """
    result = {}
    while True:
        if lesson is None:
            result = {'lesson': None}
            break
        lesson_split_by_spaces = lesson.split(' ')
        if lesson_split_by_spaces[0] == 'Физическая':
            lesson_split_by_enters = lesson.split('\n')
            teachers = (
                lesson_split_by_enters[1] + lesson_split_by_enters[2]).split(',')
            result = {'lesson_name': ''.join(
                lesson_split_by_enters[0]), 'lesson_teacher': teachers}
            break
        if '\n' in lesson:
            lesson = lesson.replace('\n', ' ')
        lesson = re.sub("\s{2,}", " ", lesson)
        if '+' in lesson:
            lesson = lesson.replace('+', ' + ')
            lesson_split_by_plus = lesson.split(' + ')
            lesson_split_by_plus = list(
                map(lambda s: s.strip(), lesson_split_by_plus))
            lesson_name_and_lesson_type = lesson_split_by_plus[0]
            if '(' not in lesson:
                lesson_name = lesson_name_and_lesson_type[0]
                lesson_type = ''
            else:
                
                lesson_name = lesson_name_and_lesson_type.split('(')[0]
                lesson_type = lesson_name_and_lesson_type.split('(')[1].split(')')[
                    0]
            lesson_teacher_and_faculty = lesson_split_by_plus[1]
            lesson_name = f'{lesson_name.strip()} + {lesson_teacher_and_faculty.split(" ")[0]}'
            lesson_teacher = ' '.join(
                lesson_teacher_and_faculty.split(' ')[1:])
            result = {'lesson_name': lesson_name.strip(), 'lesson_teacher': lesson_teacher.strip(),
                      'lesson_type': lesson_type.strip()}
            
            break

        if '(' not in lesson:
            if '\n' not in lesson:
                print('lesson=',lesson)
                space_indexies = [m.start() for m in re.finditer(' ', lesson)]
                result = {'lesson_name': lesson[:space_indexies[-3] + 1],
                          'lesson_teacher': lesson[space_indexies[-3] + 1:]}
                break
            else:
                split_lesson = lesson.split(' ')
                result = {'lesson_name': split_lesson[0].strip(
                ), 'lesson_teacher': split_lesson[1].strip()}
                break

        split_lesson = lesson.split('(')
        if len(split_lesson) > 2:
            indx = lesson.rfind('(')
            lesson_name = lesson[:indx]
            split_lesson = lesson[indx + 1:].split(')')
            lesson_type = split_lesson[0]
            lesson_teacher = split_lesson[1]
            result = {'lesson_name': lesson_name.strip(), 'lesson_teacher': lesson_teacher.strip(),
                      'lesson_type': lesson_type.strip()}
            break
        lesson_name = split_lesson[0]
        split_lesson = split_lesson[1].split(')')
        lesson_type = split_lesson[0]
        lesson_teacher = split_lesson[1]
        if '\n' in lesson_teacher:
            tmp = lesson_teacher.split('\n')
            lesson_teacher = tmp[1]
            lesson_name += tmp[0]
        result = {'lesson_name': lesson_name.strip(), 'lesson_teacher': lesson_teacher.strip(),
                  'lesson_type': lesson_type.strip()}
        break
    return result


def excel_to_json(filename):
    """sumary_line
    Take excel file with schedule and make from it list
    Keyword arguments:
    filename -- name of excel file
    Return list of all lessons grabed from excel file
    """

    wb = load_workbook(filename=filename)
    ws = wb.active
    schedule = {}
    course_numbers = {'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5}

    c = 0
    n = 4
    while True:

        if ws[get_column_letter(n) + '3'].value is None:
            break
        n += 3

    for i in range(4, n, 3):
        current_group = ws[get_column_letter(i) + '3'].value.split('\n')
        if ws[get_column_letter(i) + '2'].value is not None:
            current_group = f'{course_numbers[ws[get_column_letter(i) + "2"].value.strip()]}/{current_group[0]} {current_group[1]}'
        else:
            tmp = i
            while True:
                if tmp < 4:
                    break
                if ws[get_column_letter(tmp) + '2'].value is not None:
                    current_group = f'{course_numbers[ws[get_column_letter(tmp) + "2"].value.strip()]}/{current_group[0]} {current_group[1]}'
                    break
                tmp -= 3
        schedule[current_group] = []
        for j in range(4, 74):
            tmp = i - 1
            tmp_lesson = ''
            if i > 4 and ws[get_column_letter(i) + str(j)].value is None and type(
                    ws[get_column_letter(tmp) + str(j)]).__name__ == 'MergedCell' and j % 2 == 0:
                while True:
                    tmp -= 3
                    if tmp < 1:
                        break
                    if ws[get_column_letter(tmp + 1) + str(j)].value is not None:
                        tmp_lesson = lesson_to_dict(
                            ws[get_column_letter(tmp + 1) + str(j)].value)
                        break
                if type(ws[get_column_letter(i) + str(j + 1)]).__name__ == 'MergedCell':
                    schedule[current_group].append(tmp_lesson)
                else:
                    schedule[current_group].append([{
                        'numerator': True,
                        **tmp_lesson
                    },
                        {'denominator': True, **lesson_to_dict(
                            ws[get_column_letter(i) + str(j + 1)].value)}
                    ])
                continue

            if type(ws[get_column_letter(i) + str(j)]).__name__ == 'MergedCell':
                continue
            else:
                if j % 2 == 0:
                    if ws[get_column_letter(i) + str(j + 1)].value is None and type(
                            ws[get_column_letter(i) + str(j + 1)]).__name__ == 'MergedCell':
                        tmp = i - 1
                        if i > 4 and ws[get_column_letter(i) + str(j + 1)].value is None and type(
                                ws[get_column_letter(tmp) + str(j + 1)]).__name__ == 'MergedCell':
                            while True:
                                tmp -= 3
                                if tmp < 1:
                                    break
                                if ws[get_column_letter(tmp + 1) + str(j + 1)].value is not None:
                                    schedule[current_group].append([
                                        {'numerator': True, **lesson_to_dict(
                                            ws[get_column_letter(i) + str(j)].value)},
                                        {'denominator': True, **lesson_to_dict(
                                            ws[get_column_letter(tmp + 1) + str(j + 1)].value)}
                                    ])
                                    break
                            continue
                        if type(ws[get_column_letter(i + 1) + str(j)]).__name__ != 'MergedCell':
                            schedule[current_group].append([
                                {'first_group': True, **
                                    lesson_to_dict(ws[get_column_letter(i) + str(j)].value)},
                                {'second_group': True, **
                                    lesson_to_dict(ws[get_column_letter(i + 1) + str(j)].value)}
                            ])
                        else:
                            schedule[current_group].append(lesson_to_dict(
                                ws[get_column_letter(i) + str(j)].value))
                    else:
                        if type(ws[get_column_letter(i + 1) + str(j)]).__name__ != 'MergedCell' and \
                        type(ws[get_column_letter(i + 1) + str(j + 1)]).__name__ != 'MergedCell' and \
                        type(ws[get_column_letter(i) + str(j + 1)]).__name__ != 'MergedCell':
                            schedule[current_group].append([
                                {'numerator': True, 'first_group': True, **
                                    lesson_to_dict(ws[get_column_letter(i) + str(j)].value)},
                                {'numerator': True, 'second_group': True, **
                                    lesson_to_dict(ws[get_column_letter(i + 1) + str(j)].value)},
                                {'denominator': True, 'first_group': True, **lesson_to_dict(
                                    ws[get_column_letter(i) + str(j + 1)].value)},
                                {'denominator': True, 'second_group': True,  **lesson_to_dict(ws[
                                    get_column_letter(
                                        i + 1) + str(
                                        j + 1)].value)}
                            ])
                        elif type(ws[get_column_letter(i + 1) + str(j)]).__name__ != 'MergedCell' and \
                        type(ws[get_column_letter(i + 1) + str(j + 1)]).__name__ == 'MergedCell':
                            schedule[current_group].append([
                                {'numerator': True, 'first_group': True, **
                                    lesson_to_dict(ws[get_column_letter(i) + str(j)].value)},
                                {'numerator': True, 'second_group': True, **
                                    lesson_to_dict(ws[get_column_letter(i + 1) + str(j)].value)},
                                {'denominator': True,  **
                                    lesson_to_dict(ws[get_column_letter(i) + str(j + 1)].value)}
                            ])
                        elif type(ws[get_column_letter(i + 1) + str(j + 1)]).__name__ != 'MergedCell'\
                         and type(ws[get_column_letter(i + 1) + str(j)]).__name__ == 'MergedCell':
                            schedule[current_group].append([
                                {'numerator': True, **
                                    lesson_to_dict(ws[get_column_letter(i) + str(j)].value)},
                                {'denominator': True, 'first_group': True, **
                                    lesson_to_dict(ws[get_column_letter(i) + str(j + 1)].value)},
                                {'denominator': True, 'second_group': True, **lesson_to_dict(ws[get_column_letter(i + 1) + str(j + 1)].value)}])
                        else:
                            schedule[current_group].append([
                                {'numerator': True, **
                                    lesson_to_dict(ws[get_column_letter(i) + str(j)].value)},
                                {'denominator': True, **
                                    lesson_to_dict(ws[get_column_letter(i) + str(j + 1)].value)}
                            ])
        schedule[current_group] = list(split(schedule[current_group], 6))
        print(schedule)
    return schedule

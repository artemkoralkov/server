from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook
import json


def split_list_by_parts(lst: list, parts: int) -> list:
    k, m = divmod(len(lst), parts)
    return [lst[i * k + min(i, m):(i + 1) * k + min(i + 1, m)] for i in range(parts)]


def get_merged_cell_value(sheet, cell) -> str:
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng) != 0 else cell.value


def is_merged_sell(cell) -> bool:
    return type(cell).__name__ == 'MergedCell'


def lesson_to_dict(lesson: str, faculty=None, group_number=None):
    if lesson is None:
        return {'lesson': None}
    elif 'СМГ' in lesson and group_number:
        if lesson.count('\n') > 1:
            lesson = lesson[::-1].replace('\n', ' ', 1)[::-1]
        lesson_name, teachers = lesson.split('\n')
        teachers = teachers.split(', ')
        smg_teacher = teachers[[teachers.index(i) for i in teachers if 'СМГ' in i][0]]
        teachers.remove(smg_teacher)
        if group_number == 4 and faculty != 'dino':
            return {
                'lesson_title': 'Физическая культура',
                'teacher_name': f'{teachers[group_number - 2]}, {smg_teacher}'
            }
        return {
            'lesson_title': 'Физическая культура',
            'teacher_name': f'{teachers[group_number - 1]}, {smg_teacher}'
        }
    else:
        positions = ('доц.', 'пр.-ст.', 'ст.пр.', 'пр.', 'проф.')
        lesson_types = (
            'ЛК', 'ПЗ/СЗ', 'СЗ', 'СЗ/ЛЗ', 'ЛК/ПЗ', 'ПЗ', ' Л К ', 'ЛК/СЗ', '  Л  К  ', 'лк', 'пр', 'лб', 'сз',
            'лк/пр', 'пз', 'лз', 'ЛЗ', 'ЛБ', 'лк, лз'
        )
        position_index = 1
        lesson_type = ''
        lesson = lesson.replace('\n', ' ')
        for position in positions:
            position_index = lesson.find(position)
            if position_index != -1:
                break
        lesson_name = lesson[:position_index].strip()
        lesson_teacher = lesson[position_index:].strip()
        if faculty == 'dino':
            left_bracket_index = lesson_teacher.rfind('(')
            right_bracket_index = lesson_teacher.rfind(')')
            if lesson_teacher[left_bracket_index + 1: right_bracket_index] in lesson_types:
                lesson_type = lesson_teacher[left_bracket_index + 1: right_bracket_index]
        else:
            left_bracket_index = lesson_name.rfind('(')
            right_bracket_index = lesson_name.rfind(')')
            if lesson_name[left_bracket_index + 1: right_bracket_index] in lesson_types:
                lesson_type = lesson_name[left_bracket_index + 1: right_bracket_index]
    return {
        'lesson_title': lesson_name.replace(f'({lesson_type})', '').strip(),
        'teacher_name': lesson_teacher.replace(f'({lesson_type})', '').strip(),
        'lesson_type': lesson_type.strip().replace(' ', '').replace('  ', '')
    }

def excel_to_json(filename: str, faculty: str) -> 'dict[str, list[dict[str, str]]]':
    wb = load_workbook(filename)
    ws = wb.active
    schedule = {}
    course_numbers = {'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5}
    max_lessons = 6
    if faculty == 'tbfb':
        lessons_start_row = 5
        lessons_end_row = 75
        groups_names_start_row = '4'
        course_number_row = '3'
    elif faculty == 'dino':
        lessons_start_row = 4
        lessons_end_row = 64
        groups_names_start_row = '3'
        course_number_row = '2'
    else:
        lessons_start_row = 4
        lessons_end_row = 74
        groups_names_start_row = '3'
        course_number_row = '2'
    column = 4
    while ws[get_column_letter(column) + groups_names_start_row].value:
        group_number, speciality =\
            list(map(lambda e: e.strip(), ws[get_column_letter(column) + groups_names_start_row].value.split('\n')))
        course_number = get_merged_cell_value(ws, ws[get_column_letter(column) + course_number_row])
        current_group = f'{course_numbers[course_number]}/{group_number} {speciality}'
        schedule[current_group] = []
        for row in range(lessons_start_row, lessons_end_row, 2):
            upper_left_cell = ws[get_column_letter(column) + str(row)]
            upper_right_cell = ws[get_column_letter(column + 1) + str(row)]
            bottom_left_cell = ws[get_column_letter(column) + str(row + 1)]
            bottom_right_cell = ws[get_column_letter(column + 1) + str(row + 1)]
            if is_merged_sell(upper_right_cell) and \
                    is_merged_sell(bottom_left_cell) and \
                    is_merged_sell(bottom_right_cell):
                if is_merged_sell(upper_left_cell):
                    if get_merged_cell_value(ws, upper_left_cell) != get_merged_cell_value(ws, bottom_left_cell):
                        schedule[current_group].append([
                            {
                                'numerator': True,
                                **lesson_to_dict(get_merged_cell_value(ws, upper_left_cell), faculty)
                            },
                            {
                                'denominator': True,
                                **lesson_to_dict(get_merged_cell_value(ws, bottom_left_cell), faculty)
                            },
                        ])
                    else:
                        schedule[current_group].append(
                            lesson_to_dict(get_merged_cell_value(ws, upper_left_cell), faculty, int(group_number[0]))
                        )
                elif get_merged_cell_value(ws, upper_left_cell) != get_merged_cell_value(ws, bottom_left_cell):
                    schedule[current_group].append([
                        {
                            'numerator': True,
                            **lesson_to_dict(upper_left_cell.value, faculty)
                        },
                        {
                            'denominator': True,
                            **lesson_to_dict(get_merged_cell_value(ws, bottom_left_cell), faculty)
                        },
                    ])
                else:
                    schedule[current_group].append(
                        lesson_to_dict(upper_left_cell.value, faculty, int(group_number[0]))
                    )
            elif not is_merged_sell(upper_right_cell) and \
                    not is_merged_sell(bottom_left_cell) and \
                    not is_merged_sell(bottom_right_cell):
                schedule[current_group].append([
                    {
                        'numerator': True,
                        'first_group': True,
                        **lesson_to_dict(upper_left_cell.value, faculty)
                    },
                    {
                        'numerator': True,
                        'second_group': True,
                        **lesson_to_dict(upper_right_cell.value, faculty)
                    },
                    {
                        'denominator': True,
                        'first_group': True,
                        **lesson_to_dict(bottom_left_cell.value, faculty)
                    },
                    {
                        'denominator': True,
                        'second_group': True,
                        **lesson_to_dict(bottom_right_cell.value, faculty)
                    },
                ])
            elif not is_merged_sell(upper_left_cell) and \
                    is_merged_sell(bottom_right_cell):
                if not is_merged_sell(upper_right_cell) and \
                        is_merged_sell(bottom_left_cell):
                    schedule[current_group].append([
                        {
                            'first_group': True,
                            **lesson_to_dict(upper_left_cell.value, faculty)
                        },
                        {
                            'second_group': True,
                            **lesson_to_dict(upper_right_cell.value, faculty)
                        },
                    ])
                elif not is_merged_sell(upper_right_cell) and \
                        not is_merged_sell(bottom_left_cell) and \
                        get_merged_cell_value(ws, bottom_right_cell) == upper_right_cell.value:
                    schedule[current_group].append([
                        {
                            'first_group': True,
                            'numerator': True,
                            **lesson_to_dict(upper_left_cell.value, faculty)
                        },
                        {
                            'first_group': True,
                            'denominator': True,
                            **lesson_to_dict(bottom_left_cell.value, faculty)
                        },
                        {
                            'second_group': True,
                            **lesson_to_dict(upper_right_cell.value, faculty)
                        },
                    ])
                elif is_merged_sell(upper_right_cell) and \
                        not is_merged_sell(bottom_left_cell):
                    schedule[current_group].append([
                        {
                            'numerator': True,
                            **lesson_to_dict(upper_left_cell.value, faculty)
                        },
                        {
                            'denominator': True,
                            **lesson_to_dict(bottom_left_cell.value, faculty)
                        },
                    ])
                elif not is_merged_sell(upper_right_cell):
                    schedule[current_group].append([
                        {
                            'numerator': True,
                            'first_group': True,
                            **lesson_to_dict(upper_left_cell.value, faculty)
                        },
                        {
                            'numerator': True,
                            'second_group': True,
                            **lesson_to_dict(upper_right_cell.value, faculty)
                        },
                        {
                            'denominator': True,
                            **lesson_to_dict(bottom_left_cell.value, faculty)
                        },
                    ])
            elif is_merged_sell(upper_right_cell) and \
                    not is_merged_sell(bottom_left_cell) and \
                    not is_merged_sell(bottom_right_cell):
                schedule[current_group].append([
                    {
                        'numerator': True,
                        **lesson_to_dict(upper_left_cell.value, faculty)
                    },
                    {
                        'denominator': True,
                        'first_group': True,
                        **lesson_to_dict(bottom_left_cell.value, faculty)
                    },
                    {
                        'denominator': True,
                        'second_group': True,
                        **lesson_to_dict(bottom_right_cell.value, faculty)
                    },
                ])
            elif is_merged_sell(bottom_right_cell) and \
                    not is_merged_sell(upper_left_cell) and \
                    not is_merged_sell(upper_right_cell):
                schedule[current_group].append([
                    {
                        'numerator': True,
                        'first_group': True,
                        **lesson_to_dict(upper_left_cell.value, faculty)
                    },
                    {
                        'numerator': True,
                        'second_group': True,
                        **lesson_to_dict(upper_right_cell.value, faculty)
                    },
                    {
                        'denominator': True,
                        **lesson_to_dict(bottom_left_cell.value, faculty)
                    },
                ])
            elif is_merged_sell(upper_left_cell) and \
                    is_merged_sell(bottom_right_cell) and \
                    not is_merged_sell(bottom_left_cell):
                schedule[current_group].append([
                    {
                        'numerator': True,
                        **lesson_to_dict(get_merged_cell_value(ws, upper_left_cell), faculty)
                    },
                    {
                        'denominator': True,
                        **lesson_to_dict(bottom_left_cell.value, faculty)
                    },
                ])
            elif is_merged_sell(bottom_left_cell) and \
                    not is_merged_sell(upper_right_cell) and \
                    not is_merged_sell(bottom_right_cell):
                schedule[current_group].append([
                    {
                        'first_group': True,
                        **lesson_to_dict(upper_left_cell.value, faculty)
                    },
                    {
                        'numerator': True,
                        'second_group': True,
                        **lesson_to_dict(upper_right_cell.value, faculty)
                    },
                    {
                        'denominator': True,
                        'second_group': True,
                        **lesson_to_dict(bottom_right_cell.value, faculty)
                    },
                ])
            elif is_merged_sell(upper_left_cell) and is_merged_sell(bottom_left_cell) and not is_merged_sell(upper_right_cell) and is_merged_sell(bottom_right_cell):
                schedule[current_group].append([
                    {
                        'first_group': True,
                        **lesson_to_dict(get_merged_cell_value(ws, upper_left_cell), faculty)
                    },
                    {
                        'second_group': True,
                        **lesson_to_dict(upper_right_cell.value, faculty)
                    },
                ])
            else:
                print('-----')
                print(get_column_letter(column) + str(row))
                print(upper_left_cell.value, upper_right_cell.value, bottom_left_cell.value, bottom_right_cell.value)
                print('-----')
                schedule[current_group].append(None)
        column += 3
        schedule[current_group] = split_list_by_parts(schedule[current_group], max_lessons)
    return schedule


# lessons = excel_to_json("B:/Downloads/raspis_FIF__I_semestr_2022-2023.xlsx")
# with open('schedule.json', 'r', encoding='utf-8') as file:
#     schedule = json.load(file)


# print(
#     lesson_to_dict("Физическая культура\nпр. Федорович В.К., пр. Таргонский Н.Н., пр. Маслова Е.А.,\nСМГ Болбас Е.В.")
# )
# wb = load_workbook(filename)
# ws = wb.active
# lessons_start_row = 5 lessons_end_row = 75 groups_names_start_row = '4' course_number_row = '3' course_numbers = {
# 'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5} column = 4 while ws[get_column_letter(column) +
# groups_names_start_row].value: print(column) print(get_column_letter(column) + groups_names_start_row) print(list(
# map(lambda e: e.strip(), ws[get_column_letter(column) + groups_names_start_row].value.split('\n')))) group_number,
# speciality = list(map(lambda e: e.strip(), ws[get_column_letter(column) + groups_names_start_row].value.split(
# '\n'))) course_number = get_merged_cell_value(ws, ws[get_column_letter(column) + course_number_row]) current_group
# = f'{course_numbers[course_number]}/{group_number} {speciality}'
        
#         column += 3
# upper_left_cell = ws['J40']
# upper_right_cell = ws['K40']
# bottom_left_cell = ws['J41']
# bottom_right_cell = ws['K41']
# print(is_merged_sell(ws['K41']) and not is_merged_sell(ws['J40']) and not is_merged_sell(ws['K40']))
# if is_merged_sell(bottom_right_cell) and \
#         not is_merged_sell(upper_left_cell) and \
#         not is_merged_sell(upper_right_cell):
#     print([
#         {
#             'numerator': True,
#             'first_group': True,
#             **lesson_to_dict(upper_left_cell.value)
#         },
#         {
#             'numerator': True,
#             'second_group': True,
#             **lesson_to_dict(upper_right_cell.value)
#         },
#         {
#             'denominator': True,
#             **lesson_to_dict(bottom_left_cell.value)
#         },
#     ])
# filename = "F:/work/Расписания и нагрузки/2022-2023/Расписание учебных занятий факультета ДиНО (дневное) 2семестр (2022-2023).xlsx"
# with open('schedule.json', 'w', encoding='utf-8') as file:
#     json.dump(
#         excel_to_json(filename, faculty='dino'),
#         file,
#         ensure_ascii=False,
#         indent=4
#     )

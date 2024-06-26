from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

positions = (
    "доц.",
    "пр.-ст.",
    "ст.пр.",
    "ст. пр.",
    "ст. преп.",
    "ст.преп.",
    "преп.",
    "пр.",
    "проф.",
)
lesson_types = (
    "(ЛК)",
    "(ПЗ/СЗ)",
    "(СЗ)",
    "(СЗ/ЛЗ)",
    "(ЛК/ПЗ)",
    "(ПЗ)",
    "( Л К )",
    "(ЛК/СЗ)",
    "(  Л  К  )",
    "(лк)",
    "(пр)",
    "(лб)",
    "(сз)",
    "(лк/пр)",
    "(пз)",
    "(лз)",
    "(ЛЗ)",
    "(ЛБ)",
    "(лк, лз)",
    "(сз,лз)",
    "(лз, сз)",
    "(пз,лз)",
    "(лк,лз)",
    "(лк,пз)",
    "(лк, пз)",
    "(лк,пз,лз)",
    "(пз, лз)",
)


def parse_lesson_type(string: str):
    for lesson_type in lesson_types:
        if lesson_type in string:
            return lesson_type
    return ""


def parse_teacher_name(string: str) -> str:
    prev_position_index = len(string)
    for position in positions:
        position_index = string.find(position)
        if (
            position_index != -1
            and prev_position_index > position_index
            and string[position_index - 1] != "("
        ):
            prev_position_index = position_index
    teacher = string[prev_position_index:].strip()
    if "(" in teacher:
        teacher = teacher[: teacher.find("(")]
    return teacher.strip()


def parse_lesson_title(lesson: str) -> str:
    return " ".join(lesson.split())


def parse_grop_name(string: str) -> str:
    return " ".join(string.split()).strip()


def calculate_teachers_count(teachers: str) -> int:
    teachers_count = 0
    for position in positions:
        teachers_count += teachers.count(position)
        teachers = teachers.replace(position, "")
    return teachers_count


def split_list_by_parts(lst: list, parts: int) -> list:
    k, m = divmod(len(lst), parts)
    return [lst[i * k + min(i, m) : (i + 1) * k + min(i + 1, m)] for i in range(parts)]


def get_merged_cell_value(sheet, cell) -> str:
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return (
        sheet.cell(rng[0].min_row, rng[0].min_col).value
        if len(rng) != 0
        else cell.value
    )


def is_merged_sell(cell) -> bool:
    return type(cell).__name__ == "MergedCell"


def lesson_to_dict(lesson: str, faculty=None, group_number=None):
    # print(lesson)
    if lesson is None:
        return {"lesson": None}
    elif "СМГ" in lesson and group_number:
        if lesson.count("\n") > 1:
            lesson = lesson[::-1].replace("\n", " ", 1)[::-1]
        lesson_name, teachers = lesson.split("\n")
        teachers = teachers.split(", ")
        smg_teacher = teachers[[teachers.index(i) for i in teachers if "СМГ" in i][0]]
        teachers.remove(smg_teacher)
        if len(teachers) == 1:
            return {
                "lesson_title": "Физическая культура",
                "teacher_name": f"{teachers[0]}, {smg_teacher}",
            }
        if faculty == "ТБФ (технология)":
            group_number = group_number - 2
            if len(teachers) < 3:
                return {
                    "lesson_title": "Физическая культура",
                    "teacher_name": f"{', '.join(teachers)}, {smg_teacher}",
                }
        elif group_number == 4 and faculty != "ДиНО":
            return {
                "lesson_title": "Физическая культура",
                "teacher_name": f"{teachers[group_number - 2]}, {smg_teacher}",
            }
        if len(teachers) == 2 and group_number == 3:
            return {
                "lesson_title": "Физическая культура",
                "teacher_name": f"{teachers[1]}, {smg_teacher}",
            }
        return {
            "lesson_title": "Физическая культура",
            "teacher_name": f"{teachers[group_number - 1]}, {smg_teacher}",
        }
    else:
        # lesson = ' '.join(lesson.split())
        lesson_type = parse_lesson_type(lesson)
        print(lesson_type)
        lesson_without_type = lesson.replace(lesson_type, "")
        lesson_teacher = parse_teacher_name(lesson_without_type)
        lesson_title = parse_lesson_title(
            lesson_without_type.replace(lesson_teacher, "")
        )
        if calculate_teachers_count(lesson_teacher) > 1:
            if ", " in lesson_teacher:
                lesson_teacher = lesson_teacher.split(", ")
            elif "/" in lesson_teacher:
                lesson_teacher = lesson_teacher.split("/")
            elif "\n" in lesson_teacher:
                lesson_teacher = lesson_teacher.replace(",", "")
                lesson_teacher = lesson_teacher.split("\n")

        return {
            "lesson_title": lesson_title,
            "teacher_name": lesson_teacher,
            "lesson_type": lesson_type,
        }


# print(lesson_to_dict("Бел. язык (проф.лек.) (пз)                                                                                                                                                                                                                  доц.Прохоренко Л.В."))
# print(lesson_to_dict("Иностранный язык (пз)                                                                                                                             англ. яз. - преп. Коцур А.С.,                                                                       нем.яз. (ТБФ, ФИФ) - доц.Чечко Т.Н.                                               "))
# print(lesson_to_dict( "История древнерусской литературы и литературы XVIII
# века (ЛК)\nдоц. Чечко Т.Н./проф. Чайка Н.В.")) print( lesson_to_dict( "Иностранный
# язык\nст. пр. Гуцко И.Н.,\nдоц. Чечко Т.Н. (+ФФК)"))


def excel_to_json(filename: str, faculty: str):
    wb = load_workbook(filename)
    ws = wb.active
    schedule = {}
    course_numbers = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5}
    max_lessons = 6
    faculty_data = {
        "tbfb": (4, 74, "4", "3"),
        "ДиНО": (4, 64, "3", "2"),
        "ТБФ (технология)": (4, 72, "3", "2"),
        "ФФК": (5, 74, "4", "3"),
    }

    (
        lessons_start_row,
        lessons_end_row,
        groups_names_start_row,
        course_number_row,
    ) = faculty_data.get(faculty, (4, 74, "3", "2"))
    column = 4
    while ws[get_column_letter(column) + groups_names_start_row].value:
        # print(ws[get_column_letter(column) + groups_names_start_row].value.split(
        # "\n"))
        group_name = parse_grop_name(
            ws[get_column_letter(column) + groups_names_start_row].value
        )
        course_number = get_merged_cell_value(
            ws, ws[get_column_letter(column) + course_number_row]
        ).strip()
        current_group = f"{course_numbers[course_number]}/{group_name}"
        print(current_group)
        schedule[current_group] = []
        for row in range(lessons_start_row, lessons_end_row, 2):
            upper_left_cell = ws[get_column_letter(column) + str(row)]
            upper_right_cell = ws[get_column_letter(column + 1) + str(row)]
            bottom_left_cell = ws[get_column_letter(column) + str(row + 1)]
            bottom_right_cell = ws[get_column_letter(column + 1) + str(row + 1)]
            if (
                is_merged_sell(upper_right_cell)
                and is_merged_sell(bottom_left_cell)
                and is_merged_sell(bottom_right_cell)
            ):
                if is_merged_sell(upper_left_cell):
                    if get_merged_cell_value(
                        ws, upper_left_cell
                    ) != get_merged_cell_value(ws, bottom_left_cell):
                        schedule[current_group].append(
                            [
                                {
                                    "numerator": True,
                                    **lesson_to_dict(
                                        get_merged_cell_value(ws, upper_left_cell),
                                        faculty,
                                    ),
                                },
                                {
                                    "denominator": True,
                                    **lesson_to_dict(
                                        get_merged_cell_value(ws, bottom_left_cell),
                                        faculty,
                                    ),
                                },
                            ]
                        )
                    else:
                        schedule[current_group].append(
                            lesson_to_dict(
                                get_merged_cell_value(ws, upper_left_cell),
                                faculty,
                                int(group_name[0]),
                            )
                        )
                elif get_merged_cell_value(
                    ws, upper_left_cell
                ) != get_merged_cell_value(ws, bottom_left_cell):
                    schedule[current_group].append(
                        [
                            {
                                "numerator": True,
                                **lesson_to_dict(upper_left_cell.value, faculty),
                            },
                            {
                                "denominator": True,
                                **lesson_to_dict(
                                    get_merged_cell_value(ws, bottom_left_cell), faculty
                                ),
                            },
                        ]
                    )
                else:
                    schedule[current_group].append(
                        lesson_to_dict(
                            upper_left_cell.value, faculty, int(group_name[0])
                        )
                    )
            elif (
                not is_merged_sell(upper_right_cell)
                and not is_merged_sell(bottom_left_cell)
                and not is_merged_sell(bottom_right_cell)
            ):
                schedule[current_group].append(
                    [
                        {
                            "numerator": True,
                            "first_group": True,
                            **lesson_to_dict(upper_left_cell.value, faculty),
                        },
                        {
                            "numerator": True,
                            "second_group": True,
                            **lesson_to_dict(upper_right_cell.value, faculty),
                        },
                        {
                            "denominator": True,
                            "first_group": True,
                            **lesson_to_dict(bottom_left_cell.value, faculty),
                        },
                        {
                            "denominator": True,
                            "second_group": True,
                            **lesson_to_dict(bottom_right_cell.value, faculty),
                        },
                    ]
                )
            elif not is_merged_sell(upper_left_cell) and is_merged_sell(
                bottom_right_cell
            ):
                if not is_merged_sell(upper_right_cell) and is_merged_sell(
                    bottom_left_cell
                ):
                    schedule[current_group].append(
                        [
                            {
                                "first_group": True,
                                **lesson_to_dict(upper_left_cell.value, faculty),
                            },
                            {
                                "second_group": True,
                                **lesson_to_dict(upper_right_cell.value, faculty),
                            },
                        ]
                    )
                elif (
                    not is_merged_sell(upper_right_cell)
                    and not is_merged_sell(bottom_left_cell)
                    and get_merged_cell_value(ws, bottom_right_cell)
                    == upper_right_cell.value
                ):
                    schedule[current_group].append(
                        [
                            {
                                "first_group": True,
                                "numerator": True,
                                **lesson_to_dict(upper_left_cell.value, faculty),
                            },
                            {
                                "first_group": True,
                                "denominator": True,
                                **lesson_to_dict(bottom_left_cell.value, faculty),
                            },
                            {
                                "second_group": True,
                                **lesson_to_dict(upper_right_cell.value, faculty),
                            },
                        ]
                    )
                elif is_merged_sell(upper_right_cell) and not is_merged_sell(
                    bottom_left_cell
                ):
                    schedule[current_group].append(
                        [
                            {
                                "numerator": True,
                                **lesson_to_dict(upper_left_cell.value, faculty),
                            },
                            {
                                "denominator": True,
                                **lesson_to_dict(bottom_left_cell.value, faculty),
                            },
                        ]
                    )
                elif not is_merged_sell(upper_right_cell):
                    schedule[current_group].append(
                        [
                            {
                                "numerator": True,
                                "first_group": True,
                                **lesson_to_dict(upper_left_cell.value, faculty),
                            },
                            {
                                "numerator": True,
                                "second_group": True,
                                **lesson_to_dict(upper_right_cell.value, faculty),
                            },
                            {
                                "denominator": True,
                                **lesson_to_dict(bottom_left_cell.value, faculty),
                            },
                        ]
                    )
            elif (
                is_merged_sell(upper_right_cell)
                and not is_merged_sell(bottom_left_cell)
                and not is_merged_sell(bottom_right_cell)
            ):
                schedule[current_group].append(
                    [
                        {
                            "numerator": True,
                            **lesson_to_dict(upper_left_cell.value, faculty),
                        },
                        {
                            "denominator": True,
                            "first_group": True,
                            **lesson_to_dict(bottom_left_cell.value, faculty),
                        },
                        {
                            "denominator": True,
                            "second_group": True,
                            **lesson_to_dict(bottom_right_cell.value, faculty),
                        },
                    ]
                )
            elif (
                is_merged_sell(bottom_right_cell)
                and not is_merged_sell(upper_left_cell)
                and not is_merged_sell(upper_right_cell)
            ):
                schedule[current_group].append(
                    [
                        {
                            "numerator": True,
                            "first_group": True,
                            **lesson_to_dict(upper_left_cell.value, faculty),
                        },
                        {
                            "numerator": True,
                            "second_group": True,
                            **lesson_to_dict(upper_right_cell.value, faculty),
                        },
                        {
                            "denominator": True,
                            **lesson_to_dict(bottom_left_cell.value, faculty),
                        },
                    ]
                )
            elif (
                is_merged_sell(upper_left_cell)
                and is_merged_sell(bottom_right_cell)
                and not is_merged_sell(bottom_left_cell)
            ):
                schedule[current_group].append(
                    [
                        {
                            "numerator": True,
                            **lesson_to_dict(
                                get_merged_cell_value(ws, upper_left_cell), faculty
                            ),
                        },
                        {
                            "denominator": True,
                            **lesson_to_dict(bottom_left_cell.value, faculty),
                        },
                    ]
                )
            elif (
                is_merged_sell(bottom_left_cell)
                and not is_merged_sell(upper_right_cell)
                and not is_merged_sell(bottom_right_cell)
            ):
                schedule[current_group].append(
                    [
                        {
                            "first_group": True,
                            **lesson_to_dict(upper_left_cell.value, faculty),
                        },
                        {
                            "numerator": True,
                            "second_group": True,
                            **lesson_to_dict(upper_right_cell.value, faculty),
                        },
                        {
                            "denominator": True,
                            "second_group": True,
                            **lesson_to_dict(bottom_right_cell.value, faculty),
                        },
                    ]
                )
            elif (
                is_merged_sell(upper_left_cell)
                and is_merged_sell(bottom_left_cell)
                and not is_merged_sell(upper_right_cell)
                and is_merged_sell(bottom_right_cell)
            ):
                schedule[current_group].append(
                    [
                        {
                            "first_group": True,
                            **lesson_to_dict(
                                get_merged_cell_value(ws, upper_left_cell), faculty
                            ),
                        },
                        {
                            "second_group": True,
                            **lesson_to_dict(upper_right_cell.value, faculty),
                        },
                    ]
                )
            else:
                print("-----")
                print(get_column_letter(column) + str(row))
                print(
                    upper_left_cell.value,
                    upper_right_cell.value,
                    bottom_left_cell.value,
                    bottom_right_cell.value,
                )
                print("-----")
                schedule[current_group].append(None)
        column += 3
        if faculty == "ТБФ (технология)":
            schedule[current_group].append([{"lesson": None}])
        # print((json.dumps(schedule[current_group], ensure_ascii=False, indent=4)))
        schedule[current_group] = split_list_by_parts(
            schedule[current_group], max_lessons
        )
    return schedule


# s = excel_to_json('../../Основное 2023-2024 с 18.09.xlsx', 'ФФ')
# print(json.dumps(s, ensure_ascii=False, indent=4))

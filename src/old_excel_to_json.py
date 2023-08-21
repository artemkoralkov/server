import re

from openpyxl.utils import get_column_letter


from openpyxl.reader.excel import load_workbook


def split(a, n):
    k, m = divmod(len(a), n)
    return (a[i * k + min(i, m) : (i + 1) * k + min(i + 1, m)] for i in range(n))


def lesson_to_dict(
    lesson: str,
) -> dict[str, str] | dict[str, str | list[str]] | dict[str, None]:
    """
    Keyword arguments:
    argument -- description
    Return: return_description
    """
    if lesson is None:
        return {"lesson": None}
    lesson_split_by_spaces: list[str] = lesson.split(" ")
    if lesson_split_by_spaces[0] == "Физическая":
        lesson_split_by_enters: list[str] = lesson.split("\n")
        teachers: list[str] = (
            lesson_split_by_enters[1] + lesson_split_by_enters[2]
        ).split(",")
        return {
            "lesson_name": "".join(lesson_split_by_enters[0]),
            "lesson_teacher": teachers,
        }
    if "\n" in lesson:
        lesson = lesson.replace("\n", " ")
    lesson = re.sub("\s{2,}", " ", lesson)
    if "+" in lesson:
        lesson = lesson.replace("+", " + ")
        lesson_split_by_plus = lesson.split(" + ")
        lesson_split_by_plus = list(map(lambda s: s.strip(), lesson_split_by_plus))
        lesson_name_and_lesson_type = lesson_split_by_plus[0]
        if "(" not in lesson:
            lesson_name = lesson_name_and_lesson_type[0]
            lesson_type = ""
        else:
            lesson_name = lesson_name_and_lesson_type.split("(")[0]
            lesson_type = lesson_name_and_lesson_type.split("(")[1].split(")")[0]
        lesson_teacher_and_faculty: str = lesson_split_by_plus[1]
        lesson_name = (
            f'{lesson_name.strip()} + {lesson_teacher_and_faculty.split(" ")[0]}'
        )
        lesson_teacher = " ".join(lesson_teacher_and_faculty.split(" ")[1:])
        return {
            "lesson_name": lesson_name.strip(),
            "lesson_teacher": lesson_teacher.strip(),
            "lesson_type": lesson_type.strip(),
        }
    if "(" not in lesson:
        if "\n" not in lesson:
            space_indexes: list[int] = [m.start() for m in re.finditer(" ", lesson)]
            return {
                "lesson_name": lesson[: space_indexes[-3] + 1],
                "lesson_teacher": lesson[space_indexes[-3] + 1 :],
            }
        else:
            split_lesson = lesson.split(" ")
            return {
                "lesson_name": split_lesson[0].strip(),
                "lesson_teacher": split_lesson[1].strip(),
            }
    split_lesson = lesson.split("(")
    if len(split_lesson) > 2:
        indx: int = lesson.rfind("(")
        lesson_name = lesson[:indx]
        split_lesson = lesson[indx + 1 :].split(")")
        lesson_type = split_lesson[0]
        lesson_teacher = split_lesson[1]
        return {
            "lesson_name": lesson_name.strip(),
            "lesson_teacher": lesson_teacher.strip(),
            "lesson_type": lesson_type.strip(),
        }
    lesson_name: str = split_lesson[0]
    split_lesson: list[str] = split_lesson[1].split(")")
    lesson_type: str = split_lesson[0]
    lesson_teacher = split_lesson[1]
    if "\n" in lesson_teacher:
        tmp: list[str] = lesson_teacher.split("\n")
        lesson_teacher: str = tmp[1]
        lesson_name += tmp[0]
    return {
        "lesson_name": lesson_name.strip(),
        "lesson_teacher": lesson_teacher.strip(),
        "lesson_type": lesson_type.strip(),
    }


def excel_to_json(filename):
    """sumary_line
    Take Excel file with schedule and make from it list
    Keyword arguments:
    filename -- name of Excel file
    Return list of all lessons grabbed from Excel file
    """

    wb = load_workbook(filename=filename)
    ws = wb.active
    schedule = {}
    course_numbers = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5}

    n = 4
    while True:
        if ws[get_column_letter(n) + "3"].value is None:
            break
        n += 3

    for i in range(4, n, 3):
        current_group = ws[get_column_letter(i) + "3"].value.split("\n")
        if ws[get_column_letter(i) + "2"].value is not None:
            current_group = f'{course_numbers[ws[get_column_letter(i) + "2"].value.strip()]}/{current_group[0]} {current_group[1]}'
        else:
            tmp = i
            while True:
                if tmp < 4:
                    break
                if ws[get_column_letter(tmp) + "2"].value is not None:
                    current_group = f'{course_numbers[ws[get_column_letter(tmp) + "2"].value.strip()]}/{current_group[0]} {current_group[1]}'
                    break
                tmp -= 3
        schedule[current_group] = []
        for j in range(4, 74):
            tmp = i - 1
            tmp_lesson: dict[str, str] | dict[str, str | list[str]] | dict[
                str, None
            ] = dict()
            if (
                i > 4
                and ws[get_column_letter(i) + str(j)].value is None
                and type(ws[get_column_letter(tmp) + str(j)]).__name__ == "MergedCell"
                and j % 2 == 0
            ):
                while True:
                    tmp -= 3
                    if tmp < 1:
                        break
                    if ws[get_column_letter(tmp + 1) + str(j)].value is not None:
                        tmp_lesson = lesson_to_dict(
                            ws[get_column_letter(tmp + 1) + str(j)].value
                        )
                        break
                if type(ws[get_column_letter(i) + str(j + 1)]).__name__ == "MergedCell":
                    if (
                        type(ws[get_column_letter(tmp + 1) + str(j + 1)]).__name__
                        != "MergedCell"
                    ):
                        schedule[current_group].append(
                            [
                                {"numerator": True, **tmp_lesson},
                                {
                                    "denominator": True,
                                    **lesson_to_dict(
                                        ws[
                                            get_column_letter(tmp + 1) + str(j + 1)
                                        ].value
                                    ),
                                },
                            ]
                        )
                    else:
                        schedule[current_group].append(tmp_lesson)

                else:
                    schedule[current_group].append(
                        [
                            {"numerator": True, **tmp_lesson},
                            {
                                "denominator": True,
                                **lesson_to_dict(
                                    ws[get_column_letter(i) + str(j + 1)].value
                                ),
                            },
                        ]
                    )

                continue

            if type(ws[get_column_letter(i) + str(j)]).__name__ == "MergedCell":
                continue
            else:
                if j % 2 == 0:
                    if (
                        ws[get_column_letter(i) + str(j + 1)].value is None
                        and type(ws[get_column_letter(i) + str(j + 1)]).__name__
                        == "MergedCell"
                    ):
                        tmp = i - 1
                        if (
                            i > 4
                            and ws[get_column_letter(i) + str(j + 1)].value is None
                            and type(ws[get_column_letter(tmp) + str(j + 1)]).__name__
                            == "MergedCell"
                        ):
                            while True:
                                tmp -= 3
                                if tmp < 1:
                                    break
                                if (
                                    ws[get_column_letter(tmp + 1) + str(j + 1)].value
                                    is not None
                                ):
                                    if (
                                        type(ws[get_column_letter(i) + str(j)]).__name__
                                        != "MergedCell"
                                        and type(
                                            ws[get_column_letter(i + 1) + str(j)]
                                        ).__name__
                                        != "MergedCell"
                                    ):
                                        schedule[current_group].append(
                                            [
                                                {
                                                    "numerator": True,
                                                    "first_group": True,
                                                    **lesson_to_dict(
                                                        ws[
                                                            get_column_letter(i)
                                                            + str(j)
                                                        ].value
                                                    ),
                                                },
                                                {
                                                    "numerator": True,
                                                    "second_group": True,
                                                    **lesson_to_dict(
                                                        ws[
                                                            get_column_letter(i + 1)
                                                            + str(j)
                                                        ].value
                                                    ),
                                                },
                                                {
                                                    "denominator": True,
                                                    **lesson_to_dict(
                                                        ws[
                                                            get_column_letter(tmp + 1)
                                                            + str(j + 1)
                                                        ].value
                                                    ),
                                                },
                                            ]
                                        )
                                        break
                                    else:
                                        schedule[current_group].append(
                                            [
                                                {
                                                    "numerator": True,
                                                    **lesson_to_dict(
                                                        ws[
                                                            get_column_letter(i)
                                                            + str(j)
                                                        ].value
                                                    ),
                                                },
                                                {
                                                    "denominator": True,
                                                    **lesson_to_dict(
                                                        ws[
                                                            get_column_letter(tmp + 1)
                                                            + str(j + 1)
                                                        ].value
                                                    ),
                                                },
                                            ]
                                        )
                                        break
                            continue
                        if (
                            type(ws[get_column_letter(i + 1) + str(j)]).__name__
                            != "MergedCell"
                        ):
                            if (
                                type(ws[get_column_letter(i + 1) + str(j + 1)]).__name__
                                != "MergedCell"
                                and ws[get_column_letter(i + 1) + str(j + 1)].value
                                is not None
                            ):
                                schedule[current_group].append(
                                    [
                                        {
                                            "first_group": True,
                                            **lesson_to_dict(
                                                ws[get_column_letter(i) + str(j)].value
                                            ),
                                        },
                                        {
                                            "second_group": True,
                                            "numerator": True,
                                            **lesson_to_dict(
                                                ws[
                                                    get_column_letter(i + 1) + str(j)
                                                ].value
                                            ),
                                        },
                                        {
                                            "second_group": True,
                                            "denominator": True,
                                            **lesson_to_dict(
                                                ws[
                                                    get_column_letter(i + 1)
                                                    + str(j + 1)
                                                ].value
                                            ),
                                        },
                                    ]
                                )
                            else:
                                schedule[current_group].append(
                                    [
                                        {
                                            "first_group": True,
                                            **lesson_to_dict(
                                                ws[get_column_letter(i) + str(j)].value
                                            ),
                                        },
                                        {
                                            "second_group": True,
                                            **lesson_to_dict(
                                                ws[
                                                    get_column_letter(i + 1) + str(j)
                                                ].value
                                            ),
                                        },
                                    ]
                                )
                        else:
                            schedule[current_group].append(
                                lesson_to_dict(ws[get_column_letter(i) + str(j)].value)
                            )
                    else:
                        if (
                            type(ws[get_column_letter(i + 1) + str(j)]).__name__
                            != "MergedCell"
                            and type(ws[get_column_letter(i + 1) + str(j + 1)]).__name__
                            != "MergedCell"
                            and type(ws[get_column_letter(i) + str(j + 1)]).__name__
                            != "MergedCell"
                        ):
                            schedule[current_group].append(
                                [
                                    {
                                        "numerator": True,
                                        "first_group": True,
                                        **lesson_to_dict(
                                            ws[get_column_letter(i) + str(j)].value
                                        ),
                                    },
                                    {
                                        "numerator": True,
                                        "second_group": True,
                                        **lesson_to_dict(
                                            ws[get_column_letter(i + 1) + str(j)].value
                                        ),
                                    },
                                    {
                                        "denominator": True,
                                        "first_group": True,
                                        **lesson_to_dict(
                                            ws[get_column_letter(i) + str(j + 1)].value
                                        ),
                                    },
                                    {
                                        "denominator": True,
                                        "second_group": True,
                                        **lesson_to_dict(
                                            ws[
                                                get_column_letter(i + 1) + str(j + 1)
                                            ].value
                                        ),
                                    },
                                ]
                            )
                        elif (
                            type(ws[get_column_letter(i + 1) + str(j)]).__name__
                            != "MergedCell"
                            and type(ws[get_column_letter(i + 1) + str(j + 1)]).__name__
                            == "MergedCell"
                        ):
                            if (
                                f"{get_column_letter(i + 1) + str(j)}:{get_column_letter(i + 1) + str(j + 1)}"
                                in list(map(lambda e: str(e), ws.merged_cells.ranges))
                            ):
                                # print(get_column_letter(i) + str(j))
                                schedule[current_group].append(
                                    [
                                        {
                                            "numerator": True,
                                            "first_group": True,
                                            **lesson_to_dict(
                                                ws[get_column_letter(i) + str(j)].value
                                            ),
                                        },
                                        {
                                            "denominator": True,
                                            "first_group": True,
                                            **lesson_to_dict(
                                                ws[
                                                    get_column_letter(i) + str(j + 1)
                                                ].value
                                            ),
                                        },
                                        {
                                            "second_group": True,
                                            **lesson_to_dict(
                                                ws[
                                                    get_column_letter(i + 1) + str(j)
                                                ].value
                                            ),
                                        },
                                    ]
                                )
                            else:
                                schedule[current_group].append(
                                    [
                                        {
                                            "numerator": True,
                                            "first_group": True,
                                            **lesson_to_dict(
                                                ws[get_column_letter(i) + str(j)].value
                                            ),
                                        },
                                        {
                                            "numerator": True,
                                            "second_group": True,
                                            **lesson_to_dict(
                                                ws[
                                                    get_column_letter(i + 1) + str(j)
                                                ].value
                                            ),
                                        },
                                        {
                                            "denominator": True,
                                            **lesson_to_dict(
                                                ws[
                                                    get_column_letter(i) + str(j + 1)
                                                ].value
                                            ),
                                        },
                                    ]
                                )
                        elif (
                            type(ws[get_column_letter(i + 1) + str(j + 1)]).__name__
                            != "MergedCell"
                            and type(ws[get_column_letter(i + 1) + str(j)]).__name__
                            == "MergedCell"
                        ):
                            schedule[current_group].append(
                                [
                                    {
                                        "numerator": True,
                                        **lesson_to_dict(
                                            ws[get_column_letter(i) + str(j)].value
                                        ),
                                    },
                                    {
                                        "denominator": True,
                                        "first_group": True,
                                        **lesson_to_dict(
                                            ws[get_column_letter(i) + str(j + 1)].value
                                        ),
                                    },
                                    {
                                        "denominator": True,
                                        "second_group": True,
                                        **lesson_to_dict(
                                            ws[
                                                get_column_letter(i + 1) + str(j + 1)
                                            ].value
                                        ),
                                    },
                                ]
                            )
                        else:
                            schedule[current_group].append(
                                [
                                    {
                                        "numerator": True,
                                        **lesson_to_dict(
                                            ws[get_column_letter(i) + str(j)].value
                                        ),
                                    },
                                    {
                                        "denominator": True,
                                        **lesson_to_dict(
                                            ws[get_column_letter(i) + str(j + 1)].value
                                        ),
                                    },
                                ]
                            )
        schedule[current_group] = list(split(schedule[current_group], 6))
    # print()
    return schedule


# with open('shcedule.json', 'w', encoding='utf-8') as fp:
#     json.dump(excel_to_json("E:\work\Копия raspis_FIF__I_semestr_2022-2023.xlsx"), fp, ensure_ascii=False, indent=4)
